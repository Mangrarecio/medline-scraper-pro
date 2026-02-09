import time
import threading
import pandas as pd
import json
import requests
import os
import re
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import messagebox, scrolledtext, ttk
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

class ExtractorDefinitivo:
    def __init__(self, root):
        self.root = root
        self.root.title("Extractor Médico Profesional v10.0")
        self.root.geometry("950x750")
        self.root.configure(bg="#f8fafc")

        # --- INTERFAZ PROFESIONAL CLARA ---
        main_frame = tk.Frame(root, bg="#f8fafc")
        main_frame.pack(fill="both", expand=True, padx=30, pady=20)

        tk.Label(main_frame, text="🩺 MINERÍA DE DATOS MÉDICOS", 
                 font=("Arial", 18, "bold"), fg="#1e293b", bg="#f8fafc").pack(pady=10)

        # Input URL
        input_frame = tk.Frame(main_frame, bg="#f8fafc")
        input_frame.pack(fill="x", pady=10)
        tk.Label(input_frame, text="URL OBJETIVO:", font=("Arial", 10, "bold"), bg="#f8fafc").pack(side="left")
        self.url_input = tk.Entry(input_frame, font=("Arial", 10), bd=1, relief="solid")
        self.url_input.pack(side="left", fill="x", expand=True, padx=10)
        self.url_input.insert(0, "https://medlineplus.gov/spanish/healthtopics.html")

        # Botón
        self.btn_run = tk.Button(main_frame, text="🚀 INICIAR EXTRACCIÓN Y LIMPIEZA", 
                                 command=self.lanzar, bg="#2563eb", fg="white", 
                                 font=("Arial", 12, "bold"), height=2, cursor="hand2", bd=0)
        self.btn_run.pack(fill="x", pady=15)

        # Progreso
        self.bar = ttk.Progressbar(main_frame, length=700, mode='determinate')
        self.bar.pack(fill="x", pady=10)

        # Log
        self.log = scrolledtext.ScrolledText(main_frame, width=90, height=22, 
                                             bg="#ffffff", fg="#334155", 
                                             font=("Consolas", 10, "bold"), bd=1, relief="solid")
        self.log.pack(pady=10)

    def escribir(self, msg):
        self.log.insert(tk.END, f">> {msg}\n")
        self.log.see(tk.END)
        self.root.update_idletasks()

    def limpiar_excel_profesional(self, ruta):
        """Aplica el formato inteligente: Alineación TOP y Ancho de celdas."""
        wb = load_workbook(ruta)
        ws = wb.active
        # Configurar anchos
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 100 # Contenido amplio
        ws.column_dimensions['D'].width = 40

        for row in ws.iter_rows():
            for cell in row:
                # El truco clave: Alineación Superior para que no se agolpe
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        wb.save(ruta)

    def obtener_texto(self, url, headers):
        try:
            r = requests.get(url, headers=headers, timeout=10)
            if r.status_code == 200:
                s = BeautifulSoup(r.content, 'html.parser')
                cuerpo = s.find('div', id="topic-summary") or s.find('div', class_="main")
                if cuerpo:
                    p_list = [p.get_text(strip=True) for p in cuerpo.find_all('p') if len(p.get_text()) > 30]
                    return "\n\n".join(p_list)
            return None
        except: return None

    def lanzar(self):
        self.btn_run.config(state="disabled", bg="#94a3b8")
        threading.Thread(target=self.proceso_principal, daemon=True).start()

    def proceso_principal(self):
        base_url = "https://medlineplus.gov/spanish/healthtopics_"
        letras = "abcdefghijklmnopqrstuvw"
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/121.0.0.0'}
        datos = []
        
        try:
            self.escribir("INICIANDO BUCLE DE EXTRACCIÓN (A-Z)...")
            for i, letra in enumerate(letras):
                self.escribir(f"Procesando Letra: {letra.upper()}")
                r = requests.get(f"{base_url}{letra}.html", headers=headers)
                soup = BeautifulSoup(r.content, 'html.parser')
                
                enlaces = soup.select('section li a')
                for enlace in enlaces:
                    nombre = enlace.get_text(strip=True)
                    href = enlace.get('href', '')
                    
                    if "/spanish/" in href and "healthtopics" not in href:
                        url_tema = f"https://medlineplus.gov{href}" if href.startswith('/') else href
                        texto = self.obtener_texto(url_tema, headers)
                        if texto:
                            datos.append({
                                "Letra": letra.upper(),
                                "Tema": nombre,
                                "Contenido": texto,
                                "Fuente": url_tema
                            })
                            self.escribir(f"Capturado: {nombre}")

                self.bar["value"] = ((i + 1) / len(letras)) * 100
                time.sleep(0.1)

            if datos:
                # GUARDADO EXCEL
                nombre_xl = "DICCIONARIO_MEDICO_PRO.xlsx"
                df = pd.DataFrame(datos)
                df.to_excel(nombre_xl, index=False)
                
                # GUARDADO JSON
                with open("DICCIONARIO_MEDICO_PRO.json", 'w', encoding='utf-8') as f:
                    json.dump(datos, f, indent=4, ensure_ascii=False)
                
                self.escribir("Limpiando y formateando Excel...")
                self.limpiar_excel_profesional(nombre_xl)
                
                self.escribir("✅ ¡TODO LISTO! Excel y JSON generados.")
                messagebox.showinfo("Éxito", "Archivos creados y formateados correctamente.")

        except Exception as e:
            self.escribir(f"ERROR: {str(e)}")
        finally:
            self.btn_run.config(state="normal", bg="#2563eb")

if __name__ == "__main__":
    root = tk.Tk()
    ExtractorDefinitivo(root)
    root.mainloop()